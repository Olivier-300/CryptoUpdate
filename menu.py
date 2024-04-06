from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import os

def create_excel_file():
    while True:
        filename = input("Enter the name for the Excel file (without extension): ")
        filename += ".xlsx"  # Add the .xlsx extension

        if os.path.isfile(filename):
            print(f"File '{filename}' already exists. Please choose a different name.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Cryptos"
            ws['A1'] = "Symbol"
            ws['B1'] = "Token Amount"
            ws['C1'] = "New Token Amount"
            ws['D1'] = "New Value"
            ws['E1'] = "Gains/Losses (%)"

            # Bold and uppercase font style for titles
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal='center')
            for cell in ws['1:1']:  # Iterate through all cells in the first row
                cell.font = bold_font
                cell.alignment = center_align

            # Right align for all other cells
            right_align = Alignment(horizontal='right')
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = right_align

            wb.save(filename)
            print(f"Excel file '{filename}' created successfully.")
            break

def update_excel():
    print("Updating Excel...")
    # Write your update logic here
    # For demonstration, let's just print a message
    print("Excel file updated successfully.")

def main():
    while True:
        print("\nMENU:")
        print("1. Create a new Excel file.")
        print("2. Update an existing Excel file.")
        print("3. Exit")

        choice = input("Enter the number of the desired option: ")

        if choice == '1':
            create_excel_file()
        elif choice == '2':
            update_excel()
        elif choice == '3':
            print("Program terminated.")
            break
        else:
            print("Invalid choice. Please enter 1, 2, or 3.")

if __name__ == "__main__":
    main()
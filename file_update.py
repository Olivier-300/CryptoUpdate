import requests
from openpyxl import load_workbook
import psutil

def is_excel_open():
    for proc in psutil.process_iter():
        try:
            if 'EXCEL.EXE' in proc.name():
                return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    return False

def get_symbol_prices(symbol):
    url = f"https://api.binance.com/api/v3/avgPrice?symbol={symbol}"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        return {
            "symbol": symbol,
            "price": float(data["price"])
        }
    else:
        print("Error:", response.status_code)
        return None
    
def get_first_sheet_name():
    workbook = load_workbook("my_cryptos.xlsx")
    first_sheet_name = workbook.sheetnames[0]
    workbook.close()
    return first_sheet_name   
    
def get_symbols_name():
    first_sheet_name = get_first_sheet_name()

    workbook = load_workbook("my_cryptos.xlsx")
    sheet = workbook[first_sheet_name]

    last_row = sheet.max_row

    symbols_names = []
    for row in range(2, last_row + 1):
        symbol = sheet[f"A{row}"].value
        if symbol:
            symbols_names.append(symbol)

    workbook.close()
    return symbols_names

def get_new_token_amount(initial_tokens):
    symbols = get_symbols_name()
    symbol_prices = []

    for symbol in symbols:
        symbol_with_usdt = symbol + "USDT"
        price_data = get_symbol_prices(symbol_with_usdt)
        if price_data is not None:
            symbol_prices.append(price_data)
    
    new_token_amounts = []

    for i, prices in enumerate(symbol_prices):
        new_token_amount = round(initial_tokens[i] / prices["price"], 2)
        new_token_amounts.append(new_token_amount)

    return new_token_amounts

def get_new_prices(number_tokens):

    symbols = get_symbols_name()
    symbol_prices = []

    for symbol in symbols:
        symbol_with_usdt = symbol + "USDT"
        price_data = get_symbol_prices(symbol_with_usdt)
        if price_data is not None:
            symbol_prices.append(price_data)

    new_prices = []

    for i, prices in enumerate(symbol_prices):
        new_price = round(number_tokens[i] * prices["price"], 2)
        new_prices.append(new_price)

    return new_prices

def get_gains_losses(initial_price,new_value):

    gain_losses = []

    for i in range (len(new_value)):

        gain_loss = ((new_value[i] - initial_price[i]) / initial_price[i]) * 100
        gain_losses.append(f"{round(gain_loss, 5)}%")
        

    return gain_losses


def write_in_excel():

    if is_excel_open():
        print("Excel file is open. Please close it and run the program again.")
        return

    first_sheet_name = get_first_sheet_name()

    workbook = load_workbook("my_cryptos.xlsx")
    sheet = workbook[first_sheet_name]
    last_row = sheet.max_row


    initial_tokens = []
    for row in range(2, last_row + 1):
        tokens = sheet[f"B{row}"].value
        if tokens:
            initial_tokens.append(round(float(tokens), 2))

    new_token_amounts = get_new_token_amount(initial_tokens)
    # Check if C2 is already filled
    if sheet["C2"].value is not None:
        print("Column C2 (New Token Amounts) is already filled. Skipping update.")

    else:    
        for row, new_tokens in enumerate(new_token_amounts, start=2):
            sheet[f"C{row}"] = new_tokens
        print("New Token Amounts added successfully to the Excel file.")
    
    
    
    new_value = get_new_prices(new_token_amounts)

    for row, new_values in enumerate(new_value, start = 2):
        sheet[f"D{row}"] = new_values

    new_gains = get_gains_losses(initial_tokens,new_value)

    for row, new_gain in enumerate(new_gains, start=2):
        sheet[f"E{row}"]= new_gain
    
    workbook.save("my_cryptos.xlsx")
    workbook.close()
    
    return

write_in_excel()

